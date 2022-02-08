import requests
import os, glob
import os.path
import json
#import pandas as pd
import openpyxl

#global countryCode_Dic

def whois_requests(target):
    #{'whois': {'query': '104.251.122.39', 'queryType': 'IPv4', 'registry': 'ARIN', 'countryCode': 'US'}}
    #api_key발급: https://xn--c79as89aj0e29b77z.xn--3e0b707e/kor/openkey/keyCre.do"
    api_key =""
    
    targetip = open('./iplist.txt', 'r')
    default_line = [("","","")]
    #raw_data = pd.DataFrame(default_line, columns=['query','countryCode','country'])
    
    #excel work book
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.append(["IP","countryCode", "country", "orgName"])
    rownum=2
    colnum=1
    try:
        for line in targetip.readlines():
            print(rownum)
            line_sp =line.split('\n')
            URL ="http://whois.kisa.or.kr/openapi/whois.jsp?query="+line_sp[0]+"&key="+api_key+"&answer=json"
            print(URL)
            response  = requests.get(URL).json()
            print(response["whois"]["countryCode"])
            print(countryCode_Dic[response["whois"]["countryCode"]])
            sheet1.cell(row=rownum, column=colnum).value =line_sp[0]
            sheet1.cell(row=rownum, column=colnum+1).value =response["whois"]["countryCode"]
            sheet1.cell(row=rownum, column=colnum+2).value =countryCode_Dic[response["whois"]["countryCode"]]
            if response["whois"]["countryCode"] =='KR':
                sheet1.cell(row=rownum, column=colnum+3).value =response["whois"]["korean"]["ISP"]["netinfo"]["orgName"]
            #raw_data[lownum] =[response["whois"]["query"],response["whois"]["countryCode"],countryCode_Dic[response["whois"]["countryCode"]]]
            rownum =rownum+1

        #raw_data = pd.DataFrame(raw_data)
        #raw_data.to_excel(excel_writer='sample.xlsx')
    finally:
        print(response)
        wb.save('ip_query_result.xlsx')
    

def CountryCode_create():
    Country_File = open('./countryCode.txt', 'r', encoding='utf-8')
    global countryCode_Dic
    countryCode_Dic = {'00':'00'}
    for line in Country_File.readlines():
        country_list = line.split(',')
        #print(country_list[0]+":"+country_list[1])
        countryCode_Dic[country_list[0]] = country_list[1]
      
def target_ip(targetd):
    targetip = open('./iplist.txt', 'r').read().split('\n')
    files = os.listdir(targetDir)
    return files
        
if __name__ == '__main__':
    CountryCode_create()
    whois_requests("")
    print("end")
   
